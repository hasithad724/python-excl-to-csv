import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

file_path='..//input.xlsx'

# Add values to excel
workbook = load_workbook(file_path)
sheet = workbook['Day2']
sheet['A3'] = 'Sales'

# Apply color based on cell value in column A
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
for cell in sheet['A']:  # Replace 'A' with the column letter you want to format
    if cell.value == 'Sales':  # Condition to apply color
        cell.fill = yellow_fill

# Save and closed the workbook
workbook.save(file_path)
workbook.close()

# Read the Excel file
df = pd.read_excel( file_path)

# Convert the DataFrame to a CSV file
csv_file = 'automation.csv'  # The name of the output CSV file
df.to_csv(csv_file, index=False)